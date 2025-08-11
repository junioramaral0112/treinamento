import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
import plotly.express as px
import base64

# Caminho do arquivo Excel
EXCEL_FILE = r"S:\SEGURANÇA DO TRABALHO\app_treinamento\AUTORIZADOS.xlsx"
ABA_NR35 = 'NR_35'
# Esta lista agora serve como uma referência, mas o app vai ler todas as abas encontradas
OUTRAS_NRS = ['NR10', 'NR12', 'PONTE_ROLANTE', 'EMPILHADEIRA', 'AUTORIZADOS_GÁS']

# --- Funções de Carregamento de Dados ---
def carregar_dados(aba_nome):
    try:
        source_date_cols = ["DATA DE REALIZAÇÃO", "REALIZAÇÃO ASO ALTURA"]
        dtype_map = {
            "NOME": str, "UNIDADE": str, "SETOR": str,
            "ASO ALTURA": str, "NÃO POSSUI ADESIVO": str, "OBSERVAÇÃO": str
        }
        
        df = pd.read_excel(EXCEL_FILE, sheet_name=aba_nome, dtype=dtype_map)
        df.dropna(how='all', inplace=True)
        
        for col in source_date_cols:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce')

        if "DATA DE REALIZAÇÃO" in df.columns:
            df["VENCIMENTO DO TREINAMENTO"] = df["DATA DE REALIZAÇÃO"].apply(
                lambda x: x.replace(year=x.year + 2) if pd.notna(x) else pd.NaT
            )

        if "REALIZAÇÃO ASO ALTURA" in df.columns:
            df["VENCIMENTO DO ASO"] = df["REALIZAÇÃO ASO ALTURA"].apply(
                lambda x: x.replace(year=x.year + 1) if pd.notna(x) else pd.NaT
            )

        if 'OBSERVAÇÃO' not in df.columns:
            df['OBSERVAÇÃO'] = ""
        else:
            df['OBSERVAÇÃO'] = df['OBSERVAÇÃO'].astype(str).fillna('')
            
        return df
    except Exception as e:
        if "No sheet named" not in str(e):
            st.error(f"Erro ao carregar os dados da aba '{aba_nome}': {str(e)}")
        return pd.DataFrame()

def carregar_dados_outras_nrs(aba_nome):
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=aba_nome)
        df.dropna(how='all', inplace=True)
        
        date_cols = ["DATA DE REALIZAÇÃO", "VENCIMENTO DO TREINAMENTO", "VENCIMENTO DO ASO"]
        for col in date_cols:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce')
        return df
    except Exception:
        return pd.DataFrame()

# --- Funções de Manipulação do Excel ---
def sincronizar_planilha(caminho_arquivo, nome_aba, df_novo):
    try:
        df_para_salvar = df_novo.copy()
        if 'NR' in df_para_salvar.columns:
            df_para_salvar = df_para_salvar.drop(columns=['NR'])
            
        for col in df_para_salvar.columns:
            if "DATA" in col.upper() or "VENCIMENTO" in col.upper():
                df_para_salvar[col] = pd.to_datetime(df_para_salvar[col], format='%d/%m/%Y', errors='coerce')

        with pd.ExcelFile(caminho_arquivo) as xls:
            dados_existentes = {sheet_name: pd.read_excel(xls, sheet_name) for sheet_name in xls.sheet_names}
        
        dados_existentes[nome_aba] = df_para_salvar

        with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
            for nome_planilha, dataframe in dados_existentes.items():
                dataframe.to_excel(writer, sheet_name=nome_planilha, index=False)
        
        return True
    except Exception as e:
        st.error(f"Erro ao sincronizar a planilha '{nome_aba}': {e}")
        return False

def adicionar_registro_openpyxl(novo_registro, aba_nome):
    try:
        wb = load_workbook(EXCEL_FILE)
        
        if aba_nome not in wb.sheetnames:
            ws = wb.create_sheet(aba_nome)
            headers = list(novo_registro.keys())
            ws.append(headers)
        else:
            ws = wb[aba_nome]

        last_row_with_content = 0
        for row in range(1, ws.max_row + 2):
            if any(ws.cell(row=row, column=c).value is not None for c in range(1, ws.max_column + 1)):
                last_row_with_content = row
        
        proxima_linha = last_row_with_content + 1

        cabecalhos_planilha = [cell.value for cell in ws[1]]
        for col_idx, header in enumerate(cabecalhos_planilha, 1):
            ws.cell(row=proxima_linha, column=col_idx, value=novo_registro.get(header))
        
        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        st.error(f"Erro ao adicionar registro: {str(e)}")
        return False

def excluir_registro(nome, aba_nome):
    try:
        wb = load_workbook(EXCEL_FILE)
        if aba_nome not in wb.sheetnames:
            st.warning(f"A aba '{aba_nome}' não existe!")
            return False
        ws = wb[aba_nome]
        linha_para_excluir = next((row[0].row for row in ws.iter_rows(min_row=2) if row[0].value == nome), None)
        if linha_para_excluir:
            ws.delete_rows(linha_para_excluir)
            wb.save(EXCEL_FILE)
            return True
        else:
            st.warning(f"Registro '{nome}' não encontrado na aba '{aba_nome}'")
            return False
    except Exception as e:
        st.error(f"Erro ao excluir registro: {str(e)}")
        return False

def atualizar_registro(aba_nome, nome_original, dados_atualizados):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[aba_nome]
        
        headers = {cell.value: i for i, cell in enumerate(ws[1], 1)}
        col_nome_idx = headers.get("NOME")

        if not col_nome_idx:
            st.error(f"Coluna 'NOME' não encontrada na aba '{aba_nome}'.")
            return False

        linha_para_atualizar = next((i for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2) if row[col_nome_idx - 1] == nome_original), None)
            
        if linha_para_atualizar:
            if "DATA DE REALIZAÇÃO" in dados_atualizados:
                data_realizacao = pd.to_datetime(dados_atualizados["DATA DE REALIZAÇÃO"], format='%d/%m/%Y', errors='coerce')
                if pd.notna(data_realizacao):
                    dados_atualizados["VENCIMENTO DO TREINAMENTO"] = data_realizacao.replace(year=data_realizacao.year + 2)

            if "REALIZAÇÃO ASO ALTURA" in dados_atualizados:
                data_aso = pd.to_datetime(dados_atualizados["REALIZAÇÃO ASO ALTURA"], format='%d/%m/%Y', errors='coerce')
                if pd.notna(data_aso):
                    dados_atualizados["VENCIMENTO DO ASO"] = data_aso.replace(year=data_aso.year + 1)
                else:
                    dados_atualizados["VENCIMENTO DO ASO"] = None

            for col_name, value in dados_atualizados.items():
                if col_name in headers:
                    col_idx = headers[col_name]
                    if isinstance(value, str) and ("DATA" in col_name.upper() or "VENCIMENTO" in col_name.upper()):
                        value = pd.to_datetime(value, format='%d/%m/%Y', errors='coerce')
                        if pd.isna(value):
                            value = None
                    ws.cell(row=linha_para_atualizar, column=col_idx, value=value)
            
            wb.save(EXCEL_FILE)
            return True
        else:
            st.warning(f"Registro '{nome_original}' não encontrado para atualização.")
            return False
    except Exception as e:
        st.error(f"Erro crítico ao atualizar registro: {str(e)}")
        return False

@st.cache_data
def convert_df_to_csv(df):
    return df.to_csv(index=False, sep=';').encode('utf-8-sig')

# --- Funções de Interface Gráfica ---
def criar_cabecalho():
    try:
        def load_image(path):
            with open(path, "rb") as f:
                return base64.b64encode(f.read()).decode()

        page_param = st.query_params.get('page', 'nr35')
        if page_param == "outras_nrs":
            bg_img = load_image(r"S:\SEGURANÇA DO TRABALHO\app_treinamento\schaefer.png")
        else:
            bg_img = load_image(r"S:\SEGURANÇA DO TRABALHO\app_treinamento\nova510.png")
            nr_img = load_image(r"S:\SEGURANÇA DO TRABALHO\app_treinamento\nr-35.png")

        logo_img = load_image(r"S:\SEGURANÇA DO TRABALHO\app_treinamento\logo.png")
        sesmt_img = load_image(r"S:\SEGURANÇA DO TRABALHO\app_treinamento\sesmt.png")

        st.markdown(f"""
        <style>
        .header-container {{ position: relative; background-image: url("data:image/png;base64,{bg_img}"); background-size: cover; background-position: center; height: 280px; border-radius: 10px; margin-bottom: 25px; }}
        .header-text {{ position: absolute; top: 50%; left: 50%; transform: translate(-50%, -50%); color: white; font-size: 2.5rem; font-weight: bold; text-shadow: 2px 2px 4px rgba(0,0,0,0.7); }}
        .header-logo-left {{ position: absolute; top: 80px; left: 80px; height: 80px; }}
        .header-logos-right {{ position: absolute; top: 50px; right: 50px; display: flex; flex-direction: column; gap: 30px; }}
        .header-logos-right img {{ height: 70px; border-radius: 5px; }}
        .header-logo-right-single {{ position: absolute; top: 80px; right: 80px; }}
        .header-logo-right-single img {{ height: 80px; border-radius: 5px; }}
        .menu-navegacao {{ display: flex; justify-content: center; gap: 20px; margin-bottom: 20px; }}
        .menu-navegacao a {{ text-decoration: none; }}
        .menu-navegacao button {{ padding: 10px 20px; font-size: 16px; border-radius: 5px; border: none; cursor: pointer; background-color: #f0f2f6; }}
        .menu-navegacao button:hover {{ background-color: #d0d2d6; }}
        .menu-navegacao button.active {{ background-color: #4CAF50; color: white; }}
        </style>
        """, unsafe_allow_html=True)

        if page_param == "outras_nrs":
            st.markdown(f"""
            <div class="header-container">
                <img src="data:image/png;base64,{logo_img}" class="header-logo-left">
                <div class="header-logo-right-single"><img src="data:image/png;base64,{sesmt_img}"></div>
                <div class="header-text">Controle de Treinamentos NR</div>
            </div>""", unsafe_allow_html=True)
        else:
            st.markdown(f"""
            <div class="header-container">
                <img src="data:image/png;base64,{logo_img}" class="header-logo-left">
                <div class="header-logos-right">
                    <img src="data:image/png;base64,{nr_img}">
                    <img src="data:image/png;base64,{sesmt_img}">
                </div>
                <div class="header-text">Controle de Treinamentos NR</div>
            </div>""", unsafe_allow_html=True)
    except Exception as e:
        st.error(f"Erro ao carregar o cabeçalho: {str(e)}")

# --- Início da Aplicação ---
st.set_page_config(page_title="Controle NR", layout="wide")

# --- ALTERADO: Aplica a cor de fundo em toda a aplicação ---
st.markdown("""
<style>
.stApp {
    background-color: #9cf0ec;
}
</style>
""", unsafe_allow_html=True)
# --- FIM DA ALTERAÇÃO ---

criar_cabecalho()

if 'page' not in st.query_params:
    st.query_params.page = "nr35"
pagina_atual = st.query_params.page

st.markdown(f"""
<div class="menu-navegacao">
    <a href="?page=nr35" target="_self"><button class="{"active" if pagina_atual == "nr35" else ""}">NR 35</button></a>
    <a href="?page=outras_nrs" target="_self"><button class="{"active" if pagina_atual == "outras_nrs" else ""}">Outras NR's</button></a>
</div>
""", unsafe_allow_html=True)

# --- Lógica da Página NR35 ---
if pagina_atual == "nr35":
    df = carregar_dados(ABA_NR35)
    
    if df.empty:
        df = pd.DataFrame(columns=["NOME", "UNIDADE", "SETOR", "DATA DE REALIZAÇÃO", "VENCIMENTO DO TREINAMENTO", "REALIZAÇÃO ASO ALTURA", "VENCIMENTO DO ASO", "ASO ALTURA", "NÃO POSSUI ADESIVO", "OBSERVAÇÃO"])

    st.subheader("🔎 Filtros e Exportação")
    df_filtrado = df.copy() 

    col1, col2, col3 = st.columns(3)
    with col1:
        filtro_nome = st.text_input("Filtrar por Nome:")
        if filtro_nome:
            df_filtrado = df_filtrado[df_filtrado['NOME'].str.contains(filtro_nome, case=False, na=False)]

    with col2:
        setores = sorted(df['SETOR'].dropna().unique())
        filtro_setor = st.multiselect("Filtrar por Setor:", options=setores, default=[])
        if filtro_setor:
            df_filtrado = df_filtrado[df_filtrado['SETOR'].isin(filtro_setor)]

    with col3:
        if "VENCIMENTO DO TREINAMENTO" in df_filtrado.columns and not df_filtrado["VENCIMENTO DO TREINAMENTO"].isnull().all():
            data_min = df_filtrado["VENCIMENTO DO TREINAMENTO"].min().date()
            data_max = df_filtrado["VENCIMENTO DO TREINAMENTO"].max().date()

            filtro_data_vencimento = st.date_input(
                "Filtrar por Vencimento do Treinamento:",
                value=(),
                min_value=data_min,
                max_value=data_max,
                format="DD/MM/YYYY"
            )
            if len(filtro_data_vencimento) == 2:
                start_date, end_date = filtro_data_vencimento
                df_filtrado = df_filtrado[
                    (df_filtrado['VENCIMENTO DO TREINAMENTO'].dt.date >= start_date) &
                    (df_filtrado['VENCIMENTO DO TREINAMENTO'].dt.date <= end_date)
                ]
        else:
            st.text("Sem datas de vencimento para filtrar.")

    csv = convert_df_to_csv(df_filtrado)
    st.download_button(
        label="📥 Exportar para CSV",
        data=csv,
        file_name=f'export_nr35_filtrado_{datetime.now().strftime("%Y%m%d")}.csv',
        mime='text/csv',
    )
    
    df_display = df_filtrado.copy()
    date_cols_display = ["DATA DE REALIZAÇÃO", "VENCIMENTO DO TREINAMENTO", "REALIZAÇÃO ASO ALTURA", "VENCIMENTO DO ASO"]
    for col in date_cols_display:
        if col in df_display.columns:
            df_display[col] = pd.to_datetime(df_display[col], errors='coerce').dt.strftime("%d/%m/%Y").fillna("")
    
    st.subheader("📋 Tabela de Registros (Edite os campos necessários):")
    
    with st.form("editar_registros_form"):
        edited_df = st.data_editor(
            df_display, 
            column_config={
                "VENCIMENTO DO TREINAMENTO": st.column_config.TextColumn(disabled=True),
                "VENCIMENTO DO ASO": st.column_config.TextColumn(disabled=True),
            },
            hide_index=True,
            use_container_width=True
        )
        
        if st.form_submit_button("Salvar Alterações"):
            try:
                diff = edited_df.compare(df_display)
                if not diff.empty:
                    for idx in diff.index.unique():
                        if idx < len(df_filtrado):
                            nome_original = df_filtrado.iloc[idx]["NOME"]
                            dados_atualizados = edited_df.iloc[idx].to_dict()
                            if atualizar_registro(ABA_NR35, nome_original, dados_atualizados):
                                st.success(f"Registro de '{nome_original}' atualizado com sucesso!")
                    st.rerun()
                else:
                    st.info("Nenhuma alteração foi detectada.")
            except Exception as e:
                st.error(f"Erro ao processar alterações: {str(e)}")

    hoje = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    
    def verificar_status(data_vencimento):
        if pd.isna(data_vencimento): return "Sem Data"
        if data_vencimento < hoje: return "Vencido"
        if data_vencimento <= hoje + timedelta(days=30): return "Vencendo"
        return "OK"

    df["Status Treinamento"] = df["VENCIMENTO DO TREINAMENTO"].apply(verificar_status)
    df["Status ASO"] = df["VENCIMENTO DO ASO"].apply(verificar_status)
    
    st.write("---")
    with st.container():
        color_map = {"Vencido": "#FF5252", "Vencendo": "#FFA726", "OK": "#66BB6A", "Sem Data": "#B0BEC5"}
        st.subheader("📊 Status de Vencimento - Detalhado:")
        col1, col2 = st.columns(2)
        
        with col1:
            treinamento_counts = df["Status Treinamento"].value_counts().reindex(["OK", "Vencendo", "Vencido", "Sem Data"]).dropna()
            if not treinamento_counts.empty:
                fig_treinamento = px.pie(values=treinamento_counts.values, names=treinamento_counts.index, title="Status do Treinamento NR 35", color=treinamento_counts.index, color_discrete_map=color_map, hole=0.3)
                st.plotly_chart(fig_treinamento, use_container_width=True)
            
            st.subheader("Treinamentos Vencidos e Vencendo")
            treinamento_vencido = df[df["Status Treinamento"].isin(["Vencido", "Vencendo"])][["NOME", "UNIDADE", "SETOR", "VENCIMENTO DO TREINAMENTO", "Status Treinamento"]]
            if not treinamento_vencido.empty:
                treinamento_vencido["VENCIMENTO DO TREINAMENTO"] = treinamento_vencido["VENCIMENTO DO TREINAMENTO"].dt.strftime("%d/%m/%Y")
                def color_row_treinamento(row):
                    if row["Status Treinamento"] == "Vencido": return ['background-color: #FF5252; color: white'] * len(row)
                    elif row["Status Treinamento"] == "Vencendo": return ['background-color: #FFA726; color: white'] * len(row)
                    return [''] * len(row)
                st.dataframe(treinamento_vencido.style.apply(color_row_treinamento, axis=1), hide_index=True, use_container_width=True)

        with col2:
            aso_counts = df["Status ASO"].value_counts().reindex(["OK", "Vencendo", "Vencido", "Sem Data"]).dropna()
            if not aso_counts.empty:
                fig_aso = px.pie(values=aso_counts.values, names=aso_counts.index, title="Status do ASO para Altura", color=aso_counts.index, color_discrete_map=color_map, hole=0.3)
                st.plotly_chart(fig_aso, use_container_width=True)

            st.subheader("ASOs Vencidos, Vencendo e Sem Data")
            aso_vencido = df[df["Status ASO"].isin(["Vencido", "Vencendo", "Sem Data"])][["NOME", "UNIDADE", "SETOR", "VENCIMENTO DO ASO", "Status ASO"]]
            if not aso_vencido.empty:
                aso_vencido["VENCIMENTO DO ASO"] = aso_vencido["VENCIMENTO DO ASO"].dt.strftime("%d/%m/%Y")
                def color_row_aso(row):
                    if row["Status ASO"] == "Vencido": return ['background-color: #FF5252; color: white'] * len(row)
                    elif row["Status ASO"] == "Vencendo": return ['background-color: #FFA726; color: white'] * len(row)
                    elif row["Status ASO"] == "Sem Data": return ['background-color: #B0BEC5; color: black'] * len(row)
                    return [''] * len(row)
                st.dataframe(aso_vencido.style.apply(color_row_aso, axis=1), hide_index=True, use_container_width=True)

    with st.container():
        tab1, tab2 = st.tabs(["➕ Adicionar Registro", "➖ Excluir Registro"])
        with tab1:
            with st.form("novo_registro_form"):
                st.subheader("Adicionar Novo Registro")
                c1, c2 = st.columns(2)
                with c1:
                    nome = st.text_input("NOME*")

                    # --- ALTERADO: Campo de texto da Unidade para seleção ---
                    opcoes_unidade = sorted(['Biguaçu', 'Floripa', 'Palhoça'])
                    unidade = st.selectbox("UNIDADE*", options=opcoes_unidade, index=None, placeholder="Selecione uma unidade...")
                    
                    opcoes_setor = sorted([
                        'Acabamento', 'Astec', 'Desmolde', 'Elétrica', 'Estofaria', 
                        'Gel', 'Laminação', 'Modelagem', 'Pintura', 'Rebarba'
                    ])
                    setor = st.selectbox("SETOR*", options=opcoes_setor, index=None, placeholder="Selecione um setor...")
                    
                    aso_altura = st.selectbox("ASO ALTURA*", ["Apto", "Inapto"])
                
                with c2:
                    data_realizacao = st.date_input("Data de Realização do TREINAMENTO*", value=datetime.today())
                    data_aso_realizacao = st.date_input("Data de Realização do ASO*", value=None)
                
                nao_possui_adesivo = st.selectbox("NÃO POSSUI ADESIVO*", ["SIM", "NÃO"])
                observacao = st.text_input("OBSERVAÇÃO")

                if st.form_submit_button("Salvar"):
                    # --- ALTERADO: Validação para checar se unidade e setor foram selecionados
                    if not all([nome.strip(), unidade, setor]):
                        st.warning("Preencha todos os campos obrigatórios (*)")
                    else:
                        vencimento_treinamento = data_realizacao.replace(year=data_realizacao.year + 2) if data_realizacao else None
                        vencimento_aso = data_aso_realizacao.replace(year=data_aso_realizacao.year + 1) if data_aso_realizacao else None
                        
                        novo_registro = {
                            "NOME": nome, "UNIDADE": unidade, "SETOR": setor,
                            "DATA DE REALIZAÇÃO": datetime.combine(data_realizacao, datetime.min.time()) if data_realizacao else None,
                            "VENCIMENTO DO TREINAMENTO": datetime.combine(vencimento_treinamento, datetime.min.time()) if vencimento_treinamento else None,
                            "REALIZAÇÃO ASO ALTURA": datetime.combine(data_aso_realizacao, datetime.min.time()) if data_aso_realizacao else None,
                            "VENCIMENTO DO ASO": datetime.combine(vencimento_aso, datetime.min.time()) if vencimento_aso else None,
                            "ASO ALTURA": aso_altura,
                            "NÃO POSSUI ADESIVO": nao_possui_adesivo, "OBSERVAÇÃO": observacao
                        }
                        if adicionar_registro_openpyxl(novo_registro, ABA_NR35):
                            st.success("Registro adicionado com sucesso!")
                            st.balloons()
                            st.rerun()
        with tab2:
            with st.form("excluir_registro_form"):
                st.subheader("Excluir Registro")
                if "NOME" in df.columns and not df["NOME"].dropna().empty:
                    nomes_validos = [str(n) for n in df["NOME"].dropna().unique()]
                    nomes_disponiveis = [""] + sorted(nomes_validos)
                else:
                    nomes_disponiveis = [""]
                
                nome_excluir = st.selectbox("Selecione o NOME para excluir:", nomes_disponiveis)
                if st.form_submit_button("Excluir", type="primary"):
                    if not nome_excluir:
                        st.warning("Selecione um nome para excluir.")
                    else:
                        if excluir_registro(nome_excluir, ABA_NR35):
                            st.success(f"Registro de '{nome_excluir}' excluído!")
                            st.rerun()
                            
# --- Lógica da Página Outras NRs ---
elif pagina_atual == "outras_nrs":
    st.subheader("📋 Outras NR's")
    st.info("Adicione, edite ou remova linhas diretamente na tabela de cada NR. Clique em 'Salvar Alterações' para sincronizar com o arquivo Excel.")
    
    try:
        xls = pd.ExcelFile(EXCEL_FILE)
        abas_a_ignorar = [ABA_NR35, 'MENU']
        nrs_a_exibir = [aba for aba in xls.sheet_names if aba not in abas_a_ignorar]
    except FileNotFoundError:
        st.warning(f"Arquivo '{EXCEL_FILE}' não encontrado.")
        nrs_a_exibir = []

    nr_names = {
        'NR10': 'NR 10 - Segurança em Instalações e Serviços em Eletricidade',
        'NR12': 'NR 12 - Segurança no Trabalho em Máquinas e Equipamentos',
        'PONTE_ROLANTE': 'Ponte Rolante',
        'EMPILHADEIRA': 'Empilhadeira',
        'AUTORIZADOS_GÁS': 'Autorizados para Gás'
    }
    nr_images = {
        'NR10': r"S:\SEGURANÇA DO TRABALHO\app_treinamento\nr10.png",
        'NR12': r"S:\SEGURANÇA DO TRABALHO\app_treinamento\nr12.png",
        'PONTE_ROLANTE': r"S:\SEGURANÇA DO TRABALHO\app_treinamento\nr11.png",
        'EMPILHADEIRA': r"S:\SEGURANÇA DO TRABALHO\app_treinamento\emp.png",
        'AUTORIZADOS_GÁS': r"S:\SEGURANÇA DO TRABALHO\app_treinamento\logo.png"
    }

    if not nrs_a_exibir:
        st.warning("Nenhuma planilha de 'Outras NRs' encontrada no arquivo.")
    else:
        for nr in nrs_a_exibir:
            with st.container(border=True):
                display_name = nr_names.get(nr, nr)

                col_img, col_title = st.columns([1, 4])
                with col_img:
                    try:
                        with open(nr_images.get(nr, ""), "rb") as f:
                            img_base64 = base64.b64encode(f.read()).decode()
                            st.markdown(f'<img src="data:image/png;base64,{img_base64}" style="height: 80px; margin-top: 10px; margin-right: 15px;">', unsafe_allow_html=True)
                    except (FileNotFoundError, TypeError):
                        pass

                with col_title:
                    st.subheader(display_name)
                
                df_nr = carregar_dados_outras_nrs(nr)
                
                df_nr_filtrado = df_nr.copy()
                
                f_col1, f_col2, f_col3 = st.columns(3)
                with f_col1:
                    filtro_nome_nr = st.text_input("Filtrar por Nome:", key=f"nome_{nr}")
                    if filtro_nome_nr:
                        df_nr_filtrado = df_nr_filtrado[df_nr_filtrado['NOME'].str.contains(filtro_nome_nr, case=False, na=False)]

                with f_col2:
                    if 'SETOR' in df_nr_filtrado.columns:
                        setores_nr = sorted(df_nr_filtrado['SETOR'].dropna().unique())
                        filtro_setor_nr = st.multiselect("Filtrar por Setor:", options=setores_nr, key=f"setor_{nr}")
                        if filtro_setor_nr:
                            df_nr_filtrado = df_nr_filtrado[df_nr_filtrado['SETOR'].isin(filtro_setor_nr)]

                with f_col3:
                    if "VENCIMENTO DO TREINAMENTO" in df_nr_filtrado.columns and not df_nr_filtrado["VENCIMENTO DO TREINAMENTO"].isnull().all():
                        data_min_nr = df_nr_filtrado["VENCIMENTO DO TREINAMENTO"].min().date()
                        data_max_nr = df_nr_filtrado["VENCIMENTO DO TREINAMENTO"].max().date()
                        filtro_data_nr = st.date_input(
                            "Filtrar por Vencimento:",
                            value=(),
                            min_value=data_min_nr,
                            max_value=data_max_nr,
                            format="DD/MM/YYYY",
                            key=f"data_{nr}"
                        )
                        if len(filtro_data_nr) == 2:
                            start_date_nr, end_date_nr = filtro_data_nr
                            df_nr_filtrado = df_nr_filtrado[
                                (df_nr_filtrado['VENCIMENTO DO TREINAMENTO'].dt.date >= start_date_nr) &
                                (df_nr_filtrado['VENCIMENTO DO TREINAMENTO'].dt.date <= end_date_nr)
                            ]
                
                csv_nr = convert_df_to_csv(df_nr_filtrado)
                st.download_button(
                    label=f"📥 Exportar {nr} para CSV",
                    data=csv_nr,
                    file_name=f'export_{nr}_filtrado_{datetime.now().strftime("%Y%m%d")}.csv',
                    mime='text/csv',
                    key=f"download_{nr}"
                )
                
                df_nr_display = df_nr_filtrado.copy()

                for col in df_nr_display.columns:
                    if pd.api.types.is_datetime64_any_dtype(df_nr_display.get(col)):
                        df_nr_display[col] = df_nr_display.loc[:, col].dt.strftime('%d/%m/%Y').fillna("")

                with st.form(f"form_{nr}"):
                    edited_df_nr = st.data_editor(
                        df_nr_display,
                        hide_index=True,
                        use_container_width=True,
                        num_rows="dynamic",
                        key=f"editor_{nr}"
                    )

                    if st.form_submit_button(f"Salvar Alterações para {display_name}"):
                        if sincronizar_planilha(EXCEL_FILE, nr, edited_df_nr):
                            st.success(f"Alterações em {display_name} salvas com sucesso!")
                            st.rerun()
                        else:
                            st.error(f"Erro ao salvar alterações em {display_name}.")

# --- Rodapé ---
st.markdown(f"""
<div style="text-align: center; padding: 10px; font-size: 0.8rem; color: #555;">
    <p>Sistema de Controle de Treinamentos e Autorizados - v3.3 (Final)<br>
    Desenvolvido por <strong>Dilceu Amaral Junior</strong><br>
    {datetime.now().year}</p>
</div>
""", unsafe_allow_html=True)